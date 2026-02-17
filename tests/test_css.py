import sys

sys.path.append("src")

import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle

from condif2css.css import (
    CssBuilder,
    CssRulesRegistry,
    create_get_css_from_cell,
    get_border_styles_from_cell,
)


def _get_css_color(color: Color | None):
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        return f"#{rgb[-6:]}"
    return None


def test_css_builder_basic_methods():
    builder = CssBuilder(_get_css_color)

    assert builder.font_size(12) == ("font-size", "12px")
    assert builder.height(10, is_important=True) == ("height", "10px !important")
    assert builder.background_transparent() == ("background-color", "transparent")
    assert builder.font_underline() == ("text-decoration", "underline")
    assert builder.font_bold(is_important=True) == ("font-weight", "bold !important")
    assert builder.font_italic() == ("font-style", "italic")

    assert builder.text_align_horizontal("center") == ("text-align", "center")
    assert builder.text_align_vertical("top", is_important=True) == (
        "vertical-align",
        "top !important",
    )
    assert builder.text_align_horizontal(123) is None
    assert builder.text_align_vertical(None) is None


def test_css_builder_color_and_border_paths():
    builder = CssBuilder(_get_css_color)
    color = Color(rgb="00AABBCC")

    assert builder.font_color(color) == ("color", "#AABBCC")
    assert builder.background_color(color, is_important=True) == (
        "background-color",
        "#AABBCC !important",
    )
    assert builder.border(None, "left", color) is None

    dashed = builder.border("dashed", "left", color)
    assert ("border-left-style", "dashed") in dashed
    assert ("border-left-color", "#AABBCC") in dashed

    unknown = builder.border("unknown-style", "top", color, is_important=True)
    assert ("border-top-style", "solid !important") in unknown
    assert ("border-top-width", "1px !important") in unknown
    assert ("border-top-color", "#AABBCC !important") in unknown
    assert all(v != "1px; !important" for _, v in unknown)

    thick = builder.border("thick", "bottom", color)
    assert ("border-bottom-style", "solid") in thick
    assert ("border-bottom-width", "1px") in thick

    builder_no_color = CssBuilder(lambda _: None)
    no_color_border = builder_no_color.border("thin", "right", color)
    assert all(not k.endswith("-color") for k, _ in no_color_border)
    assert ("border-right-style", "solid") in no_color_border

    assert builder_no_color.font_color(color) is None
    assert builder_no_color.background_color(color) is None


def test_css_registry_is_stable_and_deduplicates():
    registry = CssRulesRegistry(prefix="cf")
    items_a = [("font-size", "10px"), ("color", "#111111")]
    items_b = [("color", "#111111"), ("font-size", "10px")]

    c1 = registry.register(items_a)
    c2 = registry.register(items_b)
    c3 = registry.register([("font-size", "12px")])

    assert c1 == c2
    assert c1 == "cf_x0000"
    assert c3 == "cf_x0001"

    rules = registry.get_rules()
    assert len(rules) == 2
    assert ".cf_x0000" in rules[0] or ".cf_x0000" in rules[1]


def test_get_border_styles_from_cell_handles_none_and_real_borders():
    builder = CssBuilder(_get_css_color)

    class Dummy:
        border = None

    assert get_border_styles_from_cell(Dummy(), builder) == []

    wb = Workbook()
    ws = wb.active
    ws["A1"].border = Border(
        left=Side(style="thin", color=Color(rgb="00FF0000")),
        top=Side(style="dashed", color=Color(rgb="0000FF00")),
    )

    styles = get_border_styles_from_cell(ws["A1"], builder, is_important=True)
    assert ("border-left-style", "solid !important") in styles
    assert ("border-left-width", "1px !important") in styles
    assert ("border-left-color", "#FF0000 !important") in styles
    assert ("border-top-style", "dashed !important") in styles
    assert ("border-top-color", "#00FF00 !important") in styles


def test_create_get_css_from_cell_for_regular_cell_and_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws["A1"].alignment = Alignment(horizontal="center", vertical="top")
    ws["A1"].fill = PatternFill(patternType="solid", fgColor=Color(rgb="00112233"))
    ws["A1"].font = Font(
        sz=13,
        color=Color(rgb="00DDCCBB"),
        b=True,
        i=True,
        u="single",
    )
    ws["A1"].border = Border(top=Side(style="thin", color=Color(rgb="00AA0000")))
    ws["B1"].border = Border(left=Side(style="double", color=Color(rgb="0000AA00")))

    registry = CssRulesRegistry(prefix="cf")
    builder = CssBuilder(_get_css_color)
    get_css = create_get_css_from_cell(registry, builder)

    classes = get_css(ws["A1"], merged_cell_map={"cells": [ws["B1"]]}, is_important=True)
    assert classes

    css_text = "\n".join(registry.get_rules())
    assert "text-align: center !important;" in css_text
    assert "vertical-align: top !important;" in css_text
    assert "background-color: #112233 !important;" in css_text
    assert "font-size: 13px !important;" in css_text
    assert "color: #DDCCBB !important;" in css_text
    assert "font-weight: bold !important;" in css_text
    assert "font-style: italic !important;" in css_text
    assert "text-decoration: underline !important;" in css_text
    assert "border-top-style: solid !important;" in css_text
    assert "border-left-style: double !important;" in css_text


def test_create_get_css_from_cell_differential_style_and_non_solid_warning(caplog):
    wb = Workbook()
    ws = wb.active
    ws["C1"].fill = PatternFill(patternType="gray125", fgColor=Color(rgb="00ABCDEF"))

    registry = CssRulesRegistry(prefix="cf")
    builder = CssBuilder(_get_css_color)
    get_css = create_get_css_from_cell(registry, builder)

    with caplog.at_level(logging.WARNING):
        get_css(ws["C1"])
    assert any("Pattern type is not supported" in r.message for r in caplog.records)

    dxf = DifferentialStyle(
        fill=PatternFill(bgColor=Color(rgb="00ABCDEF")),
        border=Border(bottom=Side(style="mediumDashed", color=Color(rgb="00001122"))),
        alignment=Alignment(horizontal="left", vertical="bottom"),
        font=Font(sz=9, color=Color(rgb="00010203"), b=True, i=True, u="single"),
    )
    get_css(dxf)

    css_text = "\n".join(registry.get_rules())
    assert "background-color: #ABCDEF;" in css_text
    assert "text-align: left;" in css_text
    assert "vertical-align: bottom;" in css_text
    assert "font-size: 9px;" in css_text
    assert "color: #010203;" in css_text
    assert "font-weight: bold;" in css_text
    assert "font-style: italic;" in css_text
    assert "text-decoration: underline;" in css_text
    assert "border-bottom-style: dashed;" in css_text
    assert "border-bottom-width: 2px;" in css_text
