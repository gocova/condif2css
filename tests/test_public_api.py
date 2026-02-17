import sys

sys.path.append("src")

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

import condif2css as lib


def test_public_api_exports():
    assert "get_differential_style" in lib.__all__
    assert "ThemeColorsError" in lib.__all__
    assert callable(lib.process_conditional_formatting)
    assert callable(lib.create_themed_css_color_resolver)
    assert callable(lib.get_theme_colors)
    assert callable(lib.create_get_css_from_cell)
    assert callable(lib.get_border_styles_from_cell)
    assert callable(lib.argb_to_css)
    assert issubclass(lib.ThemeColorsError, Exception)
    assert lib.CssBuilder is not None
    assert lib.CssRulesRegistry is not None


def test_get_differential_style_safe_lookup():
    wb = Workbook()
    assert lib.get_differential_style(wb, 0) is None
    assert lib.get_differential_style(wb, -1) is None
    assert lib.get_differential_style(wb, 1.5) is None  # type: ignore[arg-type]

    style = DifferentialStyle(fill=PatternFill(patternType="solid", fgColor="00FF0000"))
    wb._differential_styles.append(style)

    assert lib.get_differential_style(wb, 0) is style
    assert lib.get_differential_style(wb, 1) is None

    class DummyWorkbook:
        pass

    assert lib.get_differential_style(DummyWorkbook(), 0) is None  # type: ignore[arg-type]
