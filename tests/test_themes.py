import pytest
from openpyxl import Workbook
from openpyxl.writer.theme import theme_xml

from condif2css.themes import ThemeColorsError, get_theme_colors


def test_get_theme_colors_from_openpyxl_theme_xml():
    wb = Workbook()
    wb.loaded_theme = theme_xml

    colors = get_theme_colors(wb)

    assert len(colors) == 12
    assert colors[0] == "FFFFFF"
    assert colors[1] == "000000"
    assert colors[-1] == "800080"


def test_get_theme_colors_returns_empty_without_loaded_theme_by_default():
    wb = Workbook()
    assert get_theme_colors(wb, strict=False) == []


def test_get_theme_colors_raises_custom_exception_when_strict():
    wb = Workbook()
    with pytest.raises(ThemeColorsError):
        get_theme_colors(wb, strict=True)


def test_get_theme_colors_invalid_theme_returns_empty_or_raises():
    wb = Workbook()
    wb.loaded_theme = "<not valid xml>"

    assert get_theme_colors(wb, strict=False) == []
    with pytest.raises(ThemeColorsError):
        get_theme_colors(wb, strict=True)


def test_get_theme_colors_missing_theme_nodes_raise_in_strict_mode():
    wb = Workbook()
    wb.loaded_theme = "<a:theme xmlns:a='http://schemas.openxmlformats.org/drawingml/2006/main'/>"
    with pytest.raises(ThemeColorsError, match="themeElements"):
        get_theme_colors(wb, strict=True)


def test_get_theme_colors_missing_lastclr_raises_when_window_color():
    wb = Workbook()
    wb.loaded_theme = """
    <a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
      <a:themeElements>
        <a:clrScheme name="x">
          <a:lt1><a:sysClr val="window"/></a:lt1>
          <a:dk1><a:srgbClr val="000000"/></a:dk1>
          <a:lt2><a:srgbClr val="FFFFFF"/></a:lt2>
          <a:dk2><a:srgbClr val="111111"/></a:dk2>
          <a:accent1><a:srgbClr val="222222"/></a:accent1>
          <a:accent2><a:srgbClr val="333333"/></a:accent2>
          <a:accent3><a:srgbClr val="444444"/></a:accent3>
          <a:accent4><a:srgbClr val="555555"/></a:accent4>
          <a:accent5><a:srgbClr val="666666"/></a:accent5>
          <a:accent6><a:srgbClr val="777777"/></a:accent6>
          <a:hlink><a:srgbClr val="888888"/></a:hlink>
          <a:folHlink><a:srgbClr val="999999"/></a:folHlink>
        </a:clrScheme>
      </a:themeElements>
    </a:theme>
    """
    with pytest.raises(ThemeColorsError, match="lastClr"):
        get_theme_colors(wb, strict=True)
