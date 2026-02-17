import sys

sys.path.append("src")

from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

import condif2css.processor as processor


def _make_rule(formulas, dxf_id, priority, stop_if_true=None):
    fill = PatternFill(patternType="solid", fgColor="00FF0000")
    rule = FormulaRule(formula=formulas, fill=fill, stopIfTrue=stop_if_true)
    rule.dxfId = dxf_id
    rule.priority = priority
    return rule


def test_process_applies_expression_with_relative_offsets():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "No"
    ws["A2"] = "Yes"
    ws["A3"] = "Yes"

    ws.conditional_formatting.add(
        "A1:A3",
        _make_rule(['A1="Yes"'], dxf_id=7, priority=3, stop_if_true=True),
    )

    result = processor.process_conditional_formatting(ws)
    assert set(result.keys()) == {"Sheet1\\!A2", "Sheet1\\!A3"}
    assert result["Sheet1\\!A2"] == ("Sheet1", "A2", 3, 7, True)
    assert result["Sheet1\\!A3"] == ("Sheet1", "A3", 3, 7, True)


def test_process_keeps_lowest_priority_value_for_same_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Yes"
    ws["F1"] = "ON"

    ws.conditional_formatting.add(
        "A1",
        _make_rule(['$F$1="ON"'], dxf_id=2, priority=2),
    )
    ws.conditional_formatting.add(
        "A1",
        _make_rule(['$F$1="ON"'], dxf_id=1, priority=1),
    )

    result = processor.process_conditional_formatting(ws)
    assert result == {"Sheet1\\!A1": ("Sheet1", "A1", 1, 1, False)}


def test_process_skips_rules_without_dxf_or_with_multiple_formulas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Yes"

    ws.conditional_formatting.add(
        "A1",
        _make_rule(['A1="Yes"'], dxf_id=None, priority=1),
    )
    ws.conditional_formatting.add(
        "A1",
        _make_rule(['A1="Yes"', 'A1="No"'], dxf_id=9, priority=1),
    )

    assert processor.process_conditional_formatting(ws) == {}


def test_process_can_raise_when_fail_ok_is_false(monkeypatch):
    class ExplodingFormula:
        inputs = set()

        def __call__(self, _):
            raise RuntimeError("boom")

    monkeypatch.setattr(processor, "get_interpreter", lambda _: ExplodingFormula())

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "X"
    ws.conditional_formatting.add("A1", _make_rule(['A1="X"'], dxf_id=1, priority=1))

    with pytest.raises(RuntimeError, match="boom"):
        processor.process_conditional_formatting(ws, fail_ok=False)

    assert processor.process_conditional_formatting(ws, fail_ok=True) == {}


def test_process_returns_empty_when_conditional_formatting_is_missing():
    class DummySheet:
        conditional_formatting = None

    assert processor.process_conditional_formatting(DummySheet()) == {}


def test_get_offsets_for_variants():
    assert processor._get_offsets_for("A1", 3, 2) == (3, 2)
    assert processor._get_offsets_for("$A1", 3, 2) == (3, 0)
    assert processor._get_offsets_for("A$1", 3, 2) == (0, 2)
    assert processor._get_offsets_for("A", 3, 2) == (0, 0)


def test_extract_anchor_cell_variants():
    wb = Workbook()
    ws = wb.active
    ws["B2"] = "x"
    ws["C3"] = "y"

    assert processor._extract_anchor_cell(ws, "B2").coordinate == "B2"
    assert processor._extract_anchor_cell(ws, "B2:C3").coordinate == "B2"

    class DummySheet:
        def __getitem__(self, _):
            return object()

    assert processor._extract_anchor_cell(DummySheet(), "A1") is None


def test_process_skips_when_existing_priority_is_better():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Yes"

    ws.conditional_formatting.add(
        "A1",
        _make_rule(['A1="Yes"'], dxf_id=10, priority=1),
    )
    ws.conditional_formatting.add(
        "A1",
        _make_rule(['A1="Yes"'], dxf_id=11, priority=2),
    )

    result = processor.process_conditional_formatting(ws)
    assert result == {"Sheet1\\!A1": ("Sheet1", "A1", 1, 10, False)}


def test_process_warns_when_formula_cannot_be_parsed(monkeypatch, caplog):
    class EmptyTokenizer:
        def __init__(self, *_args, **_kwargs):
            self.items = []

    monkeypatch.setattr(processor, "Tokenizer", EmptyTokenizer)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yes"
    ws.conditional_formatting.add("A1", _make_rule(['A1="Yes"'], dxf_id=1, priority=1))

    with caplog.at_level("WARNING"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Unable to parse formula" in rec.message for rec in caplog.records)


def test_process_warns_when_interpreter_is_not_callable(monkeypatch, caplog):
    class OneTokenTokenizer:
        def __init__(self, *_args, **_kwargs):
            self.items = [SimpleNamespace(subtype="NUMBER", value="1")]

    monkeypatch.setattr(processor, "Tokenizer", OneTokenTokenizer)
    monkeypatch.setattr(processor, "get_interpreter", lambda _items: None)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yes"
    ws.conditional_formatting.add("A1", _make_rule(['A1="Yes"'], dxf_id=1, priority=1))

    with caplog.at_level("WARNING"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Unable to get callable formula" in rec.message for rec in caplog.records)


def test_process_warns_when_anchor_cell_cannot_be_extracted(monkeypatch, caplog):
    monkeypatch.setattr(processor, "_extract_anchor_cell", lambda *_args, **_kwargs: None)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yes"
    ws.conditional_formatting.add("A1", _make_rule(['A1="Yes"'], dxf_id=1, priority=1))

    with caplog.at_level("WARNING"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Unable to get anchor cell" in rec.message for rec in caplog.records)


def test_process_logs_unsupported_reference_and_non_bool_result(monkeypatch, caplog):
    class UnsupportedRefFormula:
        inputs = {"A:A"}

        def __call__(self, _):
            return True

    monkeypatch.setattr(processor, "get_interpreter", lambda _: UnsupportedRefFormula())

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yes"
    ws.conditional_formatting.add("A1", _make_rule(['A1="Yes"'], dxf_id=1, priority=1))

    with caplog.at_level("ERROR"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Unsupported reference" in rec.message for rec in caplog.records)

    class NonBoolFormula:
        inputs = set()

        def __call__(self, _):
            return "YES"

    monkeypatch.setattr(processor, "get_interpreter", lambda _: NonBoolFormula())
    with caplog.at_level("WARNING"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Expected bool for result" in rec.message for rec in caplog.records)


def test_process_handles_offset_failures(monkeypatch, caplog):
    class OffsetFormula:
        inputs = {"A1"}

        def __call__(self, _):
            return True

    def boom_offset(self, *args, **kwargs):
        raise RuntimeError("offset boom")

    monkeypatch.setattr(processor, "get_interpreter", lambda _: OffsetFormula())
    monkeypatch.setattr(processor.Cell, "offset", boom_offset)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yes"
    ws.conditional_formatting.add("A1", _make_rule(['A1="Yes"'], dxf_id=1, priority=1))

    with caplog.at_level("ERROR"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Exception while getting offset" in rec.message for rec in caplog.records)


def test_process_handles_non_cell_offset_result(monkeypatch, caplog):
    class OffsetFormula:
        inputs = {"A1"}

        def __call__(self, _):
            return True

    monkeypatch.setattr(processor, "get_interpreter", lambda _: OffsetFormula())
    monkeypatch.setattr(processor.Cell, "offset", lambda self, **kwargs: object())

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Yes"
    ws.conditional_formatting.add("A1", _make_rule(['A1="Yes"'], dxf_id=1, priority=1))

    with caplog.at_level("ERROR"):
        assert processor.process_conditional_formatting(ws) == {}
    assert any("Unable to apply 'A1'.offset" in rec.message for rec in caplog.records)


def test_iter_cells_supports_flat_iterables_of_cells():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["A2"] = "y"

    cells = list(processor._iter_cells([ws["A1"], ws["A2"]]))
    assert [cell.coordinate for cell in cells] == ["A1", "A2"]


def test_to_token_and_build_ref_values_non_set_inputs():
    assert processor._to_token(1) is not None
    assert processor._to_token(1.5) is not None
    assert processor._to_token(None) is not None
    assert processor._to_token(object()) is None

    wb = Workbook()
    ws = wb.active
    ref_values, can_apply = processor._build_ref_values(ws, ["A1"], 0, 0)
    assert ref_values == {}
    assert can_apply is True
